# ============================================================
# exporter.py — Exports cleaned DataFrame to a formatted Excel file
# Uses openpyxl for professional styling beyond basic pandas export
# ============================================================

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import logging
import os
from datetime import datetime

logger = logging.getLogger("LeadGenerator")

# ─────────────────────────────────────────────
# Style constants — tweak here to restyle easily
# ─────────────────────────────────────────────
HEADER_BG_COLOR  = "1F4E79"   # Deep navy blue
HEADER_FONT_COLOR = "FFFFFF"  # White text
ALT_ROW_COLOR    = "EBF3FB"   # Light sky blue for alternating rows
OUTPUT_FILE      = "leads.xlsx"


def apply_professional_formatting(filepath: str) -> None:
    """
    Opens the freshly exported Excel file and applies:
    - Styled header row (navy background, white bold text)
    - Alternating row colours for readability
    - Auto column widths based on content
    - Borders on all data cells
    - Centered alignment for ID and date columns
    """
    wb = load_workbook(filepath)
    ws = wb.active
    ws.title = "Leads"

    # ── Header styling ────────────────────────────────────────
    header_font   = Font(name="Arial", bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill   = PatternFill("solid", fgColor=HEADER_BG_COLOR)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:   # Row 1 = header
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # Set header row height
    ws.row_dimensions[1].height = 28

    # ── Thin border style for all data cells ─────────────────
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Alternating row colours + borders ────────────────────
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = border
            cell.font   = Font(name="Arial", size=10)
            # Apply alternating colour every other row
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Column-specific alignment ─────────────────────────────
    # Column A (Lead ID) and F (Scraped At) → centred
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_letter = cell.column_letter
            if col_letter in ("A", "F"):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # ── Auto column widths ────────────────────────────────────
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        # Clamp width: minimum 12, maximum 55
        adjusted_width = min(max(max_len + 4, 12), 55)
        ws.column_dimensions[col_letter].width = adjusted_width

    # ── Freeze top row so header stays visible while scrolling ─
    ws.freeze_panes = "A2"

    wb.save(filepath)
    logger.info("Professional formatting applied to Excel file.")


def export_to_excel(df: pd.DataFrame, filepath: str = OUTPUT_FILE) -> str:
    """
    Main export function:
    1. Writes DataFrame to Excel using pandas (fast bulk write)
    2. Applies professional openpyxl formatting on top
    3. Returns the absolute path to the saved file

    Args:
        df       : Cleaned pandas DataFrame
        filepath : Destination filename (default: leads.xlsx)

    Returns:
        Absolute path string of the saved file
    """
    try:
        logger.info(f"Exporting {len(df)} leads to {filepath} ...")

        try:
            # ── Write raw data with pandas ────────────────────────
            # index=False → don't include row numbers in the file
            df.to_excel(filepath, index=False, sheet_name="Leads", engine="openpyxl")
        except PermissionError:
            base, ext = os.path.splitext(filepath)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"{base}_{timestamp}{ext or '.xlsx'}"
            logger.warning(f"Default output file is locked. Saving to {filepath} instead.")
            print(f"   ⚠️  leads.xlsx is open or locked. Saving as {filepath} instead.")
            df.to_excel(filepath, index=False, sheet_name="Leads", engine="openpyxl")

        # ── Apply styling with openpyxl ───────────────────────
        apply_professional_formatting(filepath)

        abs_path = os.path.abspath(filepath)
        logger.info(f"Excel exported successfully → {abs_path}")
        print(f"   ✅ Excel exported successfully → {abs_path}\n")
        return abs_path

    except Exception as e:
        logger.error(f"Export failed: {e}")
        raise
